#!/usr/bin/env python3
"""Build Mellanni weekly CEO conversion/keyword report from durable CSV extracts.

Inputs and outputs must live in a durable project report directory such as:
  /media/misunderstood/DATA/projects/mellanni2025/reports/weekly_conversion/2026-06-20

This script intentionally does not call BigQuery or Gmail. It formats verified CSV
extracts into an XLSX workbook, Gmail-safe HTML body, and verification JSON.
"""
from __future__ import annotations

import argparse
import json
import math
import sys
import zipfile
from dataclasses import dataclass, asdict
from datetime import date, datetime, timedelta
from html import escape
from pathlib import Path
from typing import Any

import pandas as pd

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception as exc:  # pragma: no cover
    raise SystemExit(f"Missing openpyxl dependency: {exc}")

FORBIDDEN_OUTPUT_ROOT = Path("/home/misunderstood/temp").resolve()
DEFAULT_PROJECT_ROOT = Path("/media/misunderstood/DATA/projects/mellanni2025").resolve()

REQUIRED_FILES = {
    "validation": "validation.csv",
    "trend": "conversion_trend.csv",
    "products": "product_performance.csv",
    "asins": "asin_performance_size_color.csv",
    "keywords": "tracked_sqp_keywords.csv",
    "net_sales": "net_sales_summary.csv",
    "promos": "top_promo_discounts.csv",
}

H10_FILE = Path("h10/h10_keyword_candidates.csv")
PRIOR_H10_FILE = Path("h10/h10_prior_keyword_candidates.csv")
H10_REQUIRED_COLUMNS = {"source", "asin", "keyword", "search_volume", "organic_rank", "sponsored_rank", "cpr", "iq_score", "title_density", "competing_products", "keyword_sales", "cpc", "trend"}

REQUIRED_COLUMNS = {
    "validation": {"min_date", "max_date", "row_count", "current_rows", "prior_rows", "negative_metric_rows"},
    "trend": {"week_start", "week_end", "gross_sales", "sessions", "units", "unit_conversion"},
    "products": {"collection", "sales", "prior_sales", "sessions", "prior_sessions", "units", "prior_units", "unit_conversion", "prior_conversion", "conversion_pp"},
    "asins": {"asin", "collection", "size", "color", "sales", "prior_sales", "sessions", "prior_sessions", "units", "prior_units", "conversion", "prior_conversion", "conversion_change_pp"},
    "keywords": {"keyword", "tracked_asins", "current_position", "prior_position", "rank_movement", "search_query_volume", "total_purchases", "brand_purchases", "est_total_keyword_sales", "est_brand_keyword_sales"},
    "net_sales": {"period", "orders", "units", "gross_item_sales", "item_promo_discount", "net_item_sales", "shipping_promo_discount_excluded"},
    "promos": {"promo_label", "orders", "units", "gross_item_sales", "item_promo_discount", "net_item_sales"},
}

MONEY_COLS = {
    "gross_sales", "sales", "prior_sales", "sales_delta", "gross_item_sales", "item_promo_discount",
    "net_item_sales", "shipping_promo_discount_excluded", "est_total_keyword_sales", "est_brand_keyword_sales",
    "prior_est_total_keyword_sales",
}
PERCENT_COLS = {"unit_conversion", "prior_conversion", "conversion", "conversion_change_pp", "conversion_pp", "sales_wow", "total_sales_wow", "brand_ctr", "brand_purchase_per_click"}
INT_COLS = {"sessions", "prior_sessions", "units", "prior_units", "orders", "current_position", "prior_position", "rank_movement", "search_query_volume", "total_purchases", "brand_purchases", "search_volume", "organic_rank", "prior_organic_rank", "sponsored_rank", "title_density", "competing_products", "keyword_sales"}


@dataclass
class Check:
    name: str
    status: str
    details: str = ""


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--report-dir", type=Path, required=True, help="Durable directory containing input CSVs and receiving outputs.")
    p.add_argument("--week-end", required=True, help="Saturday week-end date, YYYY-MM-DD.")
    p.add_argument("--strict", action="store_true", help="Exit non-zero on verification warnings/failures.")
    return p.parse_args()


def fail_if_temp(path: Path) -> None:
    resolved = path.resolve()
    try:
        resolved.relative_to(FORBIDDEN_OUTPUT_ROOT)
    except ValueError:
        return
    raise SystemExit(f"Refusing to use temp path for useful report asset: {resolved}")


def week_dates(week_end_s: str) -> dict[str, date]:
    week_end = datetime.strptime(week_end_s, "%Y-%m-%d").date()
    week_start = week_end - timedelta(days=6)
    prior_end = week_start - timedelta(days=1)
    prior_start = prior_end - timedelta(days=6)
    return {"week_start": week_start, "week_end": week_end, "prior_start": prior_start, "prior_end": prior_end}


def read_inputs(report_dir: Path) -> tuple[dict[str, pd.DataFrame], list[Check]]:
    checks: list[Check] = []
    data: dict[str, pd.DataFrame] = {}
    for key, filename in REQUIRED_FILES.items():
        path = report_dir / filename
        if not path.exists():
            checks.append(Check(f"file:{filename}", "fail", "missing"))
            data[key] = pd.DataFrame()
            continue
        df = pd.read_csv(path)
        data[key] = df
        checks.append(Check(f"file:{filename}", "pass", f"rows={len(df)}"))
        missing = sorted(REQUIRED_COLUMNS[key] - set(df.columns))
        checks.append(Check(f"columns:{filename}", "fail" if missing else "pass", ", ".join(missing) if missing else "ok"))

    h10_path = report_dir / H10_FILE
    if not h10_path.exists():
        checks.append(Check("file:h10_keyword_candidates.csv", "fail", f"missing mandatory H10 artifact: {h10_path}"))
        data["h10_keywords"] = pd.DataFrame()
    else:
        h10 = pd.read_csv(h10_path)
        data["h10_keywords"] = h10
        checks.append(Check("file:h10_keyword_candidates.csv", "pass", f"rows={len(h10)}"))
        missing = sorted(H10_REQUIRED_COLUMNS - set(h10.columns))
        checks.append(Check("columns:h10_keyword_candidates.csv", "fail" if missing else "pass", ", ".join(missing) if missing else "ok"))

    prior_h10_path = report_dir / PRIOR_H10_FILE
    if prior_h10_path.exists():
        prior_h10 = pd.read_csv(prior_h10_path)
        data["prior_h10_keywords"] = prior_h10
        checks.append(Check("file:h10_prior_keyword_candidates.csv", "pass", f"rows={len(prior_h10)}"))
    else:
        data["prior_h10_keywords"] = pd.DataFrame()
        checks.append(Check("file:h10_prior_keyword_candidates.csv", "warning", "missing; H10 movement hidden because no prior H10 snapshot was provided"))
    return data, checks


def coerce_numeric(df: pd.DataFrame) -> pd.DataFrame:
    numeric_like = {"cpr", "iq_score", "cpc", "trend", "brand_ctr", "brand_purchase_per_click"}
    for col in df.columns:
        if col in MONEY_COLS or col in PERCENT_COLS or col in INT_COLS or col in numeric_like:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def fmt_money(v: Any, digits: int = 0) -> str:
    try:
        if pd.isna(v):
            return "—"
        return f"${float(v):,.{digits}f}"
    except Exception:
        return "—"


def fmt_int(v: Any) -> str:
    try:
        if pd.isna(v):
            return "—"
        return f"{float(v):,.0f}"
    except Exception:
        return "—"


def fmt_pct(v: Any, digits: int = 1, pp: bool = False) -> str:
    try:
        if pd.isna(v):
            return "—"
        val = float(v) * (100 if not pp else 100)
        suffix = " pp" if pp else "%"
        return f"{val:+.{digits}f}{suffix}" if pp else f"{val:.{digits}f}{suffix}"
    except Exception:
        return "—"


def delta_pct(current: float, prior: float) -> float | None:
    if prior is None or pd.isna(prior) or prior == 0:
        return None
    return (current - prior) / prior


def delta_line(current: float, prior: float, *, pp: bool = False) -> tuple[str, str]:
    if pp:
        d = current - prior
        color = "#0f8a4b" if d >= 0 else "#b91c1c"
        return fmt_pct(d, 2, pp=True), color
    d = delta_pct(current, prior)
    if d is None:
        return "n/a vs prior", "#667085"
    color = "#0f8a4b" if d >= 0 else "#b91c1c"
    return f"{d:+.1%} vs prior", color


def normalize_ld(label: Any) -> str:
    if label is None or pd.isna(label):
        return "LD"
    s = str(label).strip()
    if not s or s.lower() in {"nan", "none", "null", "unk"}:
        return "LD"
    return s


def format_rank_change(value: Any) -> str:
    try:
        if pd.isna(value):
            return "—"
        n = int(round(float(value)))
        if n > 0:
            return f"Up {n}"
        if n < 0:
            return f"Down {abs(n)}"
        return "Flat"
    except Exception:
        return "—"


def build_h10_keyword_table(h10: pd.DataFrame, sqp: pd.DataFrame, checks: list[Check], prior_h10: pd.DataFrame | None = None) -> pd.DataFrame:
    if h10.empty:
        checks.append(Check("h10_mandatory_rows", "fail", "No H10 rows; no H10 == no report"))
        return pd.DataFrame()
    h10 = coerce_numeric(h10.copy())
    h10 = h10[h10["keyword"].notna() & h10["keyword"].astype(str).str.strip().ne("")]
    if h10.empty:
        checks.append(Check("h10_mandatory_rows", "fail", "H10 file has no usable keyword rows"))
        return pd.DataFrame()
    checks.append(Check("h10_mandatory_rows", "pass", f"rows={len(h10)}"))
    grouped = h10.groupby("keyword", as_index=False).agg({
        "asin": lambda x: ", ".join(sorted({str(v) for v in x if pd.notna(v)})[:8]),
        "search_volume": "max",
        "organic_rank": "min",
        "sponsored_rank": "min",
        "cpr": "min",
        "iq_score": "max",
        "title_density": "min",
        "competing_products": "min",
        "keyword_sales": "max",
        "cpc": "mean",
        "trend": "max",
        "notes": "first",
    }).rename(columns={"asin": "h10_asins"})
    if prior_h10 is not None and not prior_h10.empty:
        prior_h10 = coerce_numeric(prior_h10.copy())
        prior_h10 = prior_h10[prior_h10["keyword"].notna() & prior_h10["keyword"].astype(str).str.strip().ne("")]
        prior_h10["keyword"] = prior_h10["keyword"].astype(str).str.strip().str.lower()
        prior_grouped = prior_h10.groupby("keyword", as_index=False).agg({
            "organic_rank": "min",
            "keyword_sales": "max",
        }).rename(columns={"organic_rank": "prior_organic_rank", "keyword_sales": "prior_keyword_sales"})
        grouped = grouped.merge(prior_grouped, on="keyword", how="left")
        grouped["rank_change"] = grouped["prior_organic_rank"] - grouped["organic_rank"]
        grouped.loc[grouped["prior_organic_rank"].isna() | grouped["organic_rank"].isna(), "rank_change"] = pd.NA
        grouped["rank_change_label"] = grouped["rank_change"].map(format_rank_change)
        checks.append(Check("h10_rank_movement", "pass", f"overlap_rows={int(grouped['rank_change'].notna().sum())}"))
    else:
        grouped["prior_organic_rank"] = pd.NA
        grouped["prior_keyword_sales"] = pd.NA
        grouped["rank_change"] = pd.NA
        grouped["rank_change_label"] = "—"
        checks.append(Check("h10_rank_movement", "warning", "no prior H10 snapshot; movement unavailable"))
    sqp_cols = [c for c in ["keyword", "search_query_volume", "total_purchases", "brand_purchases", "est_total_keyword_sales", "est_brand_keyword_sales", "brand_ctr", "brand_purchase_per_click"] if c in sqp.columns]
    if sqp_cols:
        grouped = grouped.merge(sqp[sqp_cols].copy(), on="keyword", how="left")
    grouped["opportunity_score"] = (
        grouped["keyword_sales"].fillna(0).rank(pct=True) * 40
        + grouped["search_volume"].fillna(0).rank(pct=True) * 25
        + grouped["iq_score"].fillna(0).rank(pct=True) * 20
        + (1 - grouped["title_density"].fillna(grouped["title_density"].max() or 0).rank(pct=True)) * 15
    ).round(1)
    grouped = grouped.sort_values(["keyword_sales", "search_volume", "opportunity_score"], ascending=False)
    overlap = grouped["est_total_keyword_sales"].notna().sum() if "est_total_keyword_sales" in grouped else 0
    checks.append(Check("h10_sqp_overlap", "pass" if overlap > 0 else "warning", f"overlap_rows={int(overlap)}"))
    return grouped


def build_context(data: dict[str, pd.DataFrame], dates: dict[str, date]) -> tuple[dict[str, Any], list[Check]]:
    checks: list[Check] = []
    for key in data:
        data[key] = coerce_numeric(data[key])

    trend = data["trend"].copy()
    net = data["net_sales"].copy()
    promos = data["promos"].copy()
    products = data["products"].copy()
    asins = data["asins"].copy()
    sqp_keywords = data["keywords"].copy()
    h10_keywords_raw = data.get("h10_keywords", pd.DataFrame()).copy()
    prior_h10_keywords_raw = data.get("prior_h10_keywords", pd.DataFrame()).copy()
    validation = data["validation"].copy()

    raw_promo_labels = promos["promo_label"].copy() if "promo_label" in promos else pd.Series(dtype=object)
    promos["promo_label"] = promos["promo_label"].map(normalize_ld)
    raw_unknown_count = raw_promo_labels.map(lambda x: normalize_ld(x) == "LD").sum() if len(raw_promo_labels) else 0
    remaining_unknown = promos["promo_label"].astype(str).str.strip().str.lower().isin({"", "nan", "none", "null", "unk"}).sum()
    checks.append(Check("promo_ld_normalization", "pass" if remaining_unknown == 0 else "fail", f"raw_unknown_labels={int(raw_unknown_count)} remaining_unknown={int(remaining_unknown)}"))

    current_trend = trend[trend["week_end"].astype(str) == str(dates["week_end"])]
    prior_trend = trend[trend["week_end"].astype(str) == str(dates["prior_end"])]
    if current_trend.empty or prior_trend.empty:
        checks.append(Check("trend_week_rows", "fail", "current/prior week rows missing"))
        cur = trend.iloc[-1].to_dict() if len(trend) else {}
        prior = trend.iloc[0].to_dict() if len(trend) else {}
    else:
        checks.append(Check("trend_week_rows", "pass", "current and prior rows present"))
        cur = current_trend.iloc[0].to_dict()
        prior = prior_trend.iloc[0].to_dict()

    cur_net = net[net["period"].astype(str) == "current"]
    prior_net = net[net["period"].astype(str) == "prior"]
    if cur_net.empty or prior_net.empty:
        checks.append(Check("net_sales_period_rows", "fail", "current/prior net rows missing"))
        cur_net_d = {}
        prior_net_d = {}
    else:
        checks.append(Check("net_sales_period_rows", "pass", "current and prior net rows present"))
        cur_net_d = cur_net.iloc[0].to_dict()
        prior_net_d = prior_net.iloc[0].to_dict()

    # Core validation checks.
    if not validation.empty:
        v = validation.iloc[0]
        checks.append(Check("business_report_days", "pass" if str(v.get("min_date")) == str(dates["prior_start"]) and str(v.get("max_date")) == str(dates["week_end"]) else "warning", f"min={v.get('min_date')} max={v.get('max_date')}"))
        checks.append(Check("negative_metrics", "pass" if float(v.get("negative_metric_rows", 1)) == 0 else "fail", f"negative_metric_rows={v.get('negative_metric_rows')}"))
    else:
        checks.append(Check("validation_row", "fail", "validation.csv empty"))

    # Reconcile Business Report gross control vs all-orders gross KPI.
    if cur and cur_net_d:
        br_gross_for_control = float(cur.get("br_gross_sales_control", cur.get("gross_sales", 0)) or 0)
        diff = abs(br_gross_for_control - float(cur_net_d.get("gross_item_sales", 0)))
        tolerance = max(1000.0, br_gross_for_control * 0.01)
        checks.append(Check("gross_reconciliation_br_vs_all_orders", "pass" if diff <= tolerance else "warning", f"diff=${diff:,.2f}; tol=${tolerance:,.2f}"))

    # Size/color labels should not be mostly numeric.
    top_asins = asins.head(25)
    numeric_size_count = top_asins.get("size", pd.Series(dtype=str)).astype(str).str.fullmatch(r"\d+(\.\d+)?").sum()
    checks.append(Check("asin_size_labels", "pass" if numeric_size_count == 0 else "warning", f"numeric_size_rows_top25={numeric_size_count}"))

    products = products.sort_values("sales", ascending=False)
    asins = asins.sort_values("sales", ascending=False)
    promos = promos.sort_values("item_promo_discount", ascending=False)
    sqp_keywords = sqp_keywords.sort_values("est_total_keyword_sales", ascending=False)
    keywords = build_h10_keyword_table(h10_keywords_raw, sqp_keywords, checks, prior_h10_keywords_raw)

    current_ao_gross = float(cur_net_d.get("gross_item_sales", 0) or 0)
    prior_ao_gross = float(prior_net_d.get("gross_item_sales", 0) or 0)
    current_br_gross = float(cur.get("br_gross_sales_control", cur.get("gross_sales", 0)) or 0)
    prior_br_gross = float(prior.get("br_gross_sales_control", prior.get("gross_sales", 0)) or 0)

    # CEO gross-sales KPI must use all-orders gross item sales when available.
    # Business Report gross is discounted/definition-shifted and remains control-only.
    metrics = {
        "gross_sales": current_ao_gross,
        "prior_gross_sales": prior_ao_gross,
        "br_gross_sales": current_br_gross,
        "prior_br_gross_sales": prior_br_gross,
        "sessions": float(cur.get("sessions", 0) or 0),
        "prior_sessions": float(prior.get("sessions", 0) or 0),
        "units": float(cur.get("units", 0) or 0),
        "prior_units": float(prior.get("units", 0) or 0),
        "conversion": float(cur.get("unit_conversion", 0) or 0),
        "prior_conversion": float(prior.get("unit_conversion", 0) or 0),
        "net_item_sales": float(cur_net_d.get("net_item_sales", 0) or 0),
        "prior_net_item_sales": float(prior_net_d.get("net_item_sales", 0) or 0),
        "item_promo_discount": float(cur_net_d.get("item_promo_discount", 0) or 0),
        "shipping_promo_discount_excluded": float(cur_net_d.get("shipping_promo_discount_excluded", 0) or 0),
        "all_orders_gross_item_sales": current_ao_gross,
    }

    if metrics["gross_sales"] and metrics["net_item_sales"]:
        checks.append(Check(
            "gross_sales_kpi_from_all_orders",
            "pass" if abs(metrics["gross_sales"] - metrics["all_orders_gross_item_sales"]) < 0.01 else "fail",
            "Gross Sales KPI uses all-orders gross_item_sales",
        ))
        checks.append(Check(
            "gross_sales_gte_net_item_sales",
            "pass" if metrics["gross_sales"] + 0.01 >= metrics["net_item_sales"] else "fail",
            f"gross={metrics['gross_sales']:,.2f}; net={metrics['net_item_sales']:,.2f}",
        ))
        expected_net = metrics["gross_sales"] - metrics["item_promo_discount"]
        checks.append(Check(
            "net_item_sales_formula",
            "pass" if abs(expected_net - metrics["net_item_sales"]) < 0.05 else "fail",
            f"gross-item_promo={expected_net:,.2f}; net={metrics['net_item_sales']:,.2f}",
        ))

    ao_gross_by_week = {
        str(dates["week_end"]): current_ao_gross,
        str(dates["prior_end"]): prior_ao_gross,
    }
    if len(trend) and current_ao_gross and prior_ao_gross:
        trend["br_gross_sales_control"] = trend["gross_sales"]
        trend["gross_sales"] = trend["week_end"].astype(str).map(ao_gross_by_week).fillna(trend["gross_sales"])

    context = {
        "dates": {k: str(v) for k, v in dates.items()},
        "metrics": metrics,
        "validation": validation,
        "trend": trend,
        "products": products,
        "asins": asins,
        "keywords": keywords,
        "sqp_keywords": sqp_keywords,
        "h10_keywords_raw": h10_keywords_raw,
        "prior_h10_keywords_raw": prior_h10_keywords_raw,
        "net_sales": net,
        "promos": promos,
    }
    return context, checks


def style_status(checks: list[Check]) -> str:
    statuses = {c.status for c in checks}
    if "fail" in statuses:
        return "fail"
    if "warning" in statuses:
        return "warning"
    return "pass"


def html_bar(value: float, max_abs: float, color: str = "#2563eb") -> str:
    if not max_abs or pd.isna(max_abs):
        pct = 0
    else:
        pct = min(100, max(0, abs(float(value)) / max_abs * 100))
    return f"<div style='background:#edf2f7;width:160px;height:10px;border-radius:6px'><div style='background:{color};width:{pct:.1f}%;height:10px;border-radius:6px'></div></div>"


def render_table(rows: list[dict[str, Any]], columns: list[tuple[str, str]], money: set[str] = set(), pct: set[str] = set(), integer: set[str] = set()) -> str:
    th = "".join(f"<th style='background:#1f2937;color:#fff;text-align:left;padding:8px;border:1px solid #e5e7eb'>{escape(label)}</th>" for key, label in columns)
    trs = []
    for row in rows:
        tds = []
        for key, label in columns:
            val = row.get(key, "")
            if key in money:
                text = fmt_money(val)
            elif key in pct:
                text = fmt_pct(val, 2, pp=True) if key.endswith("_pp") or key.endswith("change_pp") else fmt_pct(val)
            elif key in integer:
                text = fmt_int(val)
            else:
                text = escape(str(val)) if not pd.isna(val) else "—"
            color = "#242424"
            if key in {"sales_wow", "conversion_pp", "conversion_change_pp", "rank_movement"}:
                try:
                    num = float(val)
                    if key == "rank_movement":
                        color = "#0f8a4b" if num > 0 else ("#b91c1c" if num < 0 else "#667085")
                        arrow = "↓" if num > 0 else ("↑" if num < 0 else "")
                        text = f"{arrow} {abs(num):.0f}" if arrow else "0"
                    else:
                        color = "#0f8a4b" if num >= 0 else "#b91c1c"
                except Exception:
                    pass
            tds.append(f"<td style='padding:8px;border:1px solid #e5e7eb;color:{color}'>{text}</td>")
        trs.append("<tr>" + "".join(tds) + "</tr>")
    return f"<table style='border-collapse:collapse;width:100%;font-size:13px;margin:8px 0 16px'><tr>{th}</tr>{''.join(trs)}</table>"


def render_html(context: dict[str, Any]) -> str:
    d = context["dates"]
    m = context["metrics"]
    products = context["products"].head(10)
    asins = context["asins"].head(12)
    promos = context["promos"].head(10)
    keywords = context["keywords"].head(15)

    sales_delta, sales_color = delta_line(m["gross_sales"], m["prior_gross_sales"])
    net_delta, net_color = delta_line(m["net_item_sales"], m["prior_net_item_sales"])
    units_delta, units_color = delta_line(m["units"], m["prior_units"])
    conv_delta, conv_color = delta_line(m["conversion"], m["prior_conversion"], pp=True)

    kpi_cards = [
        ("Gross Sales", fmt_money(m["gross_sales"]), sales_delta, sales_color),
        ("Net Item Sales", fmt_money(m["net_item_sales"]), net_delta, net_color),
        ("Units", fmt_int(m["units"]), units_delta, units_color),
        ("Conversion", fmt_pct(m["conversion"], 2), conv_delta, conv_color),
    ]
    cards_html = "".join(
        f"<td style='background:#f8fbff;border:1px solid #dbeafe;border-radius:8px;padding:14px;width:25%;vertical-align:top'>"
        f"<div style='font-size:11px;text-transform:uppercase;color:#667085;font-weight:700'>{label}</div>"
        f"<div style='font-size:23px;font-weight:800;margin:6px 0'>{value}</div>"
        f"<div style='font-size:12px;color:{color};font-weight:700'>{delta}</div></td>"
        for label, value, delta, color in kpi_cards
    )

    top_product = products.iloc[0] if len(products) else None
    biggest_decline = context["products"].sort_values("sales_delta").iloc[0] if len(context["products"]) else None
    promo_rate = (m["item_promo_discount"] / m["all_orders_gross_item_sales"]) if m["all_orders_gross_item_sales"] else 0
    sessions_phrase = "rose to" if m["sessions"] > m["prior_sessions"] else ("fell to" if m["sessions"] < m["prior_sessions"] else "held at")
    conversion_phrase = "improved to" if m["conversion"] > m["prior_conversion"] else ("declined to" if m["conversion"] < m["prior_conversion"] else "held at")
    bullets = [
        f"Gross sales were {fmt_money(m['gross_sales'])}, {sales_delta}; net item sales were {fmt_money(m['net_item_sales'])} after {fmt_money(m['item_promo_discount'])} in item promo discounts.",
        f"Sessions {sessions_phrase} {fmt_int(m['sessions'])} while conversion {conversion_phrase} {fmt_pct(m['conversion'], 2)} ({conv_delta}).",
        f"Item promo discount rate was {fmt_pct(promo_rate, 1)}. Shipment/shipping promotions are not product promotions and are excluded from this CEO promotion view.",
    ]
    if top_product is not None:
        bullets.append(f"Top collection remained {escape(str(top_product['collection']))} at {fmt_money(top_product['sales'])} and {fmt_pct(top_product['unit_conversion'], 2)} conversion.")
    if biggest_decline is not None:
        bullets.append(f"Largest collection sales pressure: {escape(str(biggest_decline['collection']))} ({fmt_money(biggest_decline['sales_delta'])} WoW).")

    if keywords.empty:
        raise SystemExit("H10 keyword table is empty; no H10 == no report")
    h10_terms = ", ".join(escape(str(k)) for k in keywords.head(3)["keyword"].tolist())

    sales_up = m["gross_sales"] > m["prior_gross_sales"]
    sales_down = m["gross_sales"] < m["prior_gross_sales"]
    conv_up = m["conversion"] > m["prior_conversion"]
    conv_down = m["conversion"] < m["prior_conversion"]
    if sales_up and conv_down:
        first_takeaway = "Sales grew, but conversion weakened."
        first_action = "Do not simply buy more traffic. Audit offer/PDP friction first: price, coupon, inventory, reviews, hero image, and mobile content."
    elif sales_down and conv_up:
        first_takeaway = "Sales fell because traffic fell, but conversion improved."
        first_action = "Restore qualified traffic to the biggest ASINs while protecting the better conversion rate; check PPC coverage, organic rank, and lost session sources before adding more discounts."
    elif sales_down and conv_down:
        first_takeaway = "Sales and conversion both fell."
        first_action = "Fix conversion blockers before scaling traffic: review price, coupon, Buy Box, inventory, reviews, hero image, and mobile content on the top ASINs."
    elif sales_up and conv_up:
        first_takeaway = "Sales and conversion both improved."
        first_action = "Protect what is working and scale carefully: verify inventory, keep promo pressure controlled, and expand only the best-performing traffic/keyword coverage."
    else:
        first_takeaway = "Sales held roughly steady."
        first_action = "Hold the current approach, then focus this week on the largest ASIN and keyword movements instead of broad changes."
    first_why = f"Gross sales reached {fmt_money(m['gross_sales'])} ({sales_delta}); sessions {sessions_phrase} {fmt_int(m['sessions'])}; conversion {conversion_phrase} {fmt_pct(m['conversion'], 2)} ({conv_delta})."

    top_asin_details = []
    for r in asins.head(2).to_dict("records"):
        label = " ".join(str(r.get(x, "")).strip() for x in ["size", "color", "collection"] if str(r.get(x, "")).strip())
        top_asin_details.append(f"{escape(label)}: {fmt_money(r.get('sales'))}, {fmt_pct(r.get('conversion'), 2)} conv. ({fmt_pct(r.get('conversion_change_pp'), 2, pp=True)})")
    top_asin_evidence = "; ".join(top_asin_details)
    actions_html = f'''
  <h2 style="font-size:22px;margin:18px 0 10px;color:#111827">Actions / Takeaways</h2>
  <table style="width:100%;border-collapse:collapse;margin:0 0 18px;font-size:14px;line-height:1.45">
    <tr><th style="background:#7f1d1d;color:#fff;text-align:left;padding:9px;border:1px solid #fecdd3;width:25%">Takeaway</th><th style="background:#7f1d1d;color:#fff;text-align:left;padding:9px;border:1px solid #fecdd3">Why it matters</th><th style="background:#7f1d1d;color:#fff;text-align:left;padding:9px;border:1px solid #fecdd3">This week’s action</th></tr>
    <tr><td style="padding:9px;border:1px solid #fecdd3;font-weight:700">{first_takeaway}</td><td style="padding:9px;border:1px solid #fecdd3">{first_why}</td><td style="padding:9px;border:1px solid #fecdd3">{first_action}</td></tr>
    <tr><td style="padding:9px;border:1px solid #fecdd3;font-weight:700">Start with the biggest money ASINs.</td><td style="padding:9px;border:1px solid #fecdd3">{top_asin_evidence}</td><td style="padding:9px;border:1px solid #fecdd3">Assign owner to review these top ASINs and list exact fixes before next week’s report.</td></tr>
    <tr><td style="padding:9px;border:1px solid #fecdd3;font-weight:700">Use H10 for keywords; use SQP for shopper behavior.</td><td style="padding:9px;border:1px solid #fecdd3">H10 confirms priority demand pools: {h10_terms}. SQP should explain click/purchase behavior on those terms, not create a separate keyword list.</td><td style="padding:9px;border:1px solid #fecdd3">Check listing and PPC coverage for these H10 terms, then use SQP to see whether shoppers click and buy after seeing Mellanni.</td></tr>
  </table>'''

    # HTML-native charts.
    trend = context["trend"].copy()
    max_sales = trend["gross_sales"].abs().max() if len(trend) else 0
    trend_rows = "".join(
        f"<tr><td style='padding:6px'>{escape(str(r.week_end))}</td><td style='padding:6px'>{html_bar(r.gross_sales, max_sales, '#2563eb')}</td><td style='padding:6px;text-align:right'>{fmt_money(r.gross_sales)}</td><td style='padding:6px;text-align:right'>{fmt_int(r.sessions)}</td><td style='padding:6px;text-align:right'>{fmt_int(r.units)}</td><td style='padding:6px;text-align:right'>{fmt_pct(r.unit_conversion, 2)}</td></tr>"
        for r in trend.itertuples(index=False)
    )
    product_delta = context["products"].copy()
    product_delta = product_delta[product_delta["sales_delta"].abs() > 0.5]
    product_delta["abs_sales_delta"] = product_delta["sales_delta"].abs()
    product_delta = product_delta.sort_values("abs_sales_delta", ascending=False).head(8)
    max_delta = product_delta["sales_delta"].abs().max() if len(product_delta) else 0
    product_rows = "".join(
        f"<tr><td style='padding:6px'>{escape(str(r.collection))}</td><td style='padding:6px'>{html_bar(r.sales_delta, max_delta, '#b91c1c' if r.sales_delta < 0 else '#0f8a4b')}</td><td style='padding:6px;text-align:right;color:{'#b91c1c' if r.sales_delta < 0 else '#0f8a4b'}'>{fmt_money(r.sales_delta)}</td></tr>"
        for r in product_delta.itertuples(index=False)
    )
    kw_move = context["keywords"].copy().head(10)
    max_kw_sales = kw_move["keyword_sales"].max() if len(kw_move) and "keyword_sales" in kw_move else 0
    keyword_header = (
        "<tr>"
        "<th style='padding:6px;text-align:left;color:#475467;font-size:12px'>Keyword</th>"
        "<th style='padding:6px;text-align:left;color:#475467;font-size:12px'>Relative H10 keyword-sales estimate</th>"
        "<th style='padding:6px;text-align:right;color:#475467;font-size:12px'>H10 est. weekly keyword sales</th>"
        "<th style='padding:6px;text-align:right;color:#475467;font-size:12px'>Best organic rank</th>"
        "<th style='padding:6px;text-align:right;color:#475467;font-size:12px'>Rank change vs prior H10</th>"
        "</tr>"
    )
    keyword_rows = keyword_header + "".join(
        f"<tr><td style='padding:6px'>{escape(str(r.keyword))}</td><td style='padding:6px'>{html_bar(getattr(r, 'keyword_sales', 0), max_kw_sales, '#2563eb')}</td><td style='padding:6px;text-align:right'>{fmt_int(getattr(r, 'keyword_sales', 0))}</td><td style='padding:6px;text-align:right'>{fmt_int(getattr(r, 'organic_rank', 0))}</td><td style='padding:6px;text-align:right'>{escape(str(getattr(r, 'rank_change_label', '—')))}</td></tr>"
        for r in kw_move.itertuples(index=False)
    )

    product_table = render_table(products.to_dict("records"), [
        ("collection", "Collection"), ("sales", "Sales"), ("sales_wow", "Sales WoW"), ("sessions", "Sessions"), ("units", "Units"), ("unit_conversion", "Conv."), ("conversion_pp", "Conv. pp")
    ], money={"sales"}, pct={"sales_wow", "unit_conversion", "conversion_pp"}, integer={"sessions", "units"})
    asin_table = render_table(asins.to_dict("records"), [
        ("asin", "ASIN"), ("collection", "Collection"), ("size", "Size"), ("color", "Color"), ("sales", "Sales"), ("sessions", "Sessions"), ("units", "Units"), ("conversion", "Conv."), ("conversion_change_pp", "Conv. pp")
    ], money={"sales"}, pct={"conversion", "conversion_change_pp"}, integer={"sessions", "units"})
    promo_table = render_table(promos.to_dict("records"), [
        ("promo_label", "Promo"), ("item_promo_discount", "Discount"), ("gross_item_sales", "Gross"), ("net_item_sales", "Net"), ("orders", "Orders"), ("units", "Units")
    ], money={"item_promo_discount", "gross_item_sales", "net_item_sales"}, integer={"orders", "units"})
    keyword_table = render_table(keywords.to_dict("records"), [
        ("keyword", "Keyword"), ("search_volume", "H10 Search Vol."), ("keyword_sales", "H10 Kw Sales"), ("organic_rank", "Org. Rank"), ("prior_organic_rank", "Prior Org. Rank"), ("rank_change_label", "Rank Change"), ("sponsored_rank", "Spon. Rank"), ("cpr", "CPR"), ("iq_score", "IQ"), ("trend", "Trend"), ("est_brand_keyword_sales", "SQP Brand Sales"), ("brand_purchase_per_click", "Brand Purch./Click")
    ], money={"est_brand_keyword_sales"}, pct={"brand_purchase_per_click"}, integer={"search_volume", "keyword_sales", "organic_rank", "prior_organic_rank", "sponsored_rank", "trend"})

    return f"""
<div style="font-family:Arial,Helvetica,sans-serif;color:#242424;max-width:980px;margin:0 auto;background:#fff">
  <div style="background:#111827;color:#fff;padding:24px 26px;margin-bottom:14px">
    <h1 style="margin:0 0 8px;font-size:26px;line-height:1.2">Mellanni Weekly CEO Overview</h1>
    <div style="font-size:13px;color:#dbe4f0">Week ending Sat {d['week_end']} · Compared to {d['prior_start']} – {d['prior_end']}</div>
  </div>
  <div style="border-left:4px solid #f97316;background:#fff7ed;padding:11px 14px;margin:0 0 20px;font-weight:700">Prepared by Sergey's AI helper.</div>
{actions_html}

  <h2 style="font-size:20px;margin:20px 0 12px">Executive snapshot</h2>
  <table style="width:100%;border-spacing:8px;margin:0 0 12px"><tr>{cards_html}</tr></table>
  <ul style="margin:8px 0 16px;padding-left:22px;line-height:1.45">{''.join(f'<li>{b}</li>' for b in bullets)}</ul>

  <h2 style="font-size:20px;margin:20px 0 10px">Charts</h2>
  <div style="border:1px solid #d9e2ec;padding:12px;margin:8px 0 12px"><b>KPI trend</b><div style="font-size:12px;color:#667085;margin:4px 0 8px">Bars compare all-orders gross item sales within this two-week view.</div><table style="width:100%;border-collapse:collapse;font-size:13px">{trend_rows}</table></div>
  <div style="border:1px solid #d9e2ec;padding:12px;margin:8px 0 12px"><b>Collection sales delta</b><div style="font-size:12px;color:#667085;margin:4px 0 8px">Largest collection movers by absolute WoW sales delta; red = decline, green = gain.</div><table style="width:100%;border-collapse:collapse;font-size:13px">{product_rows}</table></div>
  <div style="border:1px solid #d9e2ec;padding:12px;margin:8px 0 12px"><b>H10 keyword demand / rank movement</b><div style="font-size:12px;color:#667085;margin:4px 0 8px">Bars show H10 estimated weekly keyword sales relative to the top row in this chart; rank is the best organic rank found across fetched Mellanni ASINs. Rank change compares this H10 pull against the prior report’s H10 snapshot: Up means rank improved, Down means rank fell. H10 metrics are estimates; SQP is layered only for purchase-behavior diagnostics.</div><table style="width:100%;border-collapse:collapse;font-size:13px">{keyword_rows}</table></div>

  <h2 style="font-size:18px;margin:18px 0 8px">Product / collection conversion change</h2>{product_table}
  <h2 style="font-size:18px;margin:18px 0 8px">Top ASINs</h2>{asin_table}
  <h2 style="font-size:18px;margin:18px 0 8px">Top promo discounts</h2><p style="font-size:12px;color:#667085">Net item sales deduct item promo discounts only. Shipment/shipping promotions are not product promotions and are not included in this CEO promotion view. Blank/unknown Amazon promo labels are displayed as LD.</p>{promo_table}
  <h2 style="font-size:18px;margin:18px 0 8px">H10 keyword standings with SQP diagnostics</h2><p style="font-size:12px;color:#667085">Current Helium 10/Cerebro data is the keyword source. SQP columns are behavior overlays for overlapping terms only.</p>{keyword_table}
  <p style="font-size:12px;color:#667085;margin-top:16px">Full workbook attached.</p>
</div>
""".strip()


def write_workbook(context: dict[str, Any], out: Path) -> None:
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        summary_rows = [
            {"Metric": "Week ending", "Value": context["dates"]["week_end"]},
            {"Metric": "Gross sales (all-orders)", "Value": context["metrics"]["gross_sales"]},
            {"Metric": "Net item sales (all-orders)", "Value": context["metrics"]["net_item_sales"]},
            {"Metric": "Business Report gross (control)", "Value": context["metrics"].get("br_gross_sales", 0)},
            {"Metric": "Item promo discount", "Value": context["metrics"]["item_promo_discount"]},
            {"Metric": "Shipping promo discount excluded", "Value": context["metrics"]["shipping_promo_discount_excluded"]},
            {"Metric": "Sessions", "Value": context["metrics"]["sessions"]},
            {"Metric": "Units", "Value": context["metrics"]["units"]},
            {"Metric": "Unit conversion", "Value": context["metrics"]["conversion"]},
        ]
        pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name="Summary")
        context["validation"].to_excel(writer, index=False, sheet_name="Validation")
        context["trend"].to_excel(writer, index=False, sheet_name="Trend")
        context["net_sales"].to_excel(writer, index=False, sheet_name="Net Sales")
        context["products"].to_excel(writer, index=False, sheet_name="Products")
        context["asins"].to_excel(writer, index=False, sheet_name="Top ASINs")
        context["promos"].to_excel(writer, index=False, sheet_name="Promos")
        context["keywords"].to_excel(writer, index=False, sheet_name="H10 Keywords")
        context["sqp_keywords"].to_excel(writer, index=False, sheet_name="SQP Diagnostics")

    wb = load_workbook(out)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col_cells in ws.columns:
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(max_len + 2, 10), 42)
        headers = [c.value for c in ws[1]]
        for idx, header in enumerate(headers, start=1):
            if header in MONEY_COLS or (isinstance(header, str) and any(x in header.lower() for x in ["sales", "discount"])):
                for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                    for c in cell:
                        if isinstance(c.value, (int, float)):
                            c.number_format = '$#,##0'
            if header in PERCENT_COLS or (isinstance(header, str) and any(x in header.lower() for x in ["conversion", "wow", "rate"])):
                for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                    for c in cell:
                        if isinstance(c.value, (int, float)):
                            c.number_format = '0.0%'
            if header in INT_COLS:
                for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                    for c in cell:
                        if isinstance(c.value, (int, float)):
                            c.number_format = '#,##0'
    wb.save(out)


def main() -> int:
    args = parse_args()
    report_dir = args.report_dir.resolve()
    fail_if_temp(report_dir)
    dates = week_dates(args.week_end)
    report_dir.mkdir(parents=True, exist_ok=True)

    data, checks = read_inputs(report_dir)
    context, more_checks = build_context(data, dates)
    checks.extend(more_checks)

    week_end = dates["week_end"].isoformat()
    xlsx_path = report_dir / f"Mellanni_Weekly_CEO_Overview_{week_end}.xlsx"
    html_path = report_dir / f"email_body_{week_end}.html"
    verification_path = report_dir / f"verification_summary_{week_end}.json"

    write_workbook(context, xlsx_path)
    html_path.write_text(render_html(context), encoding="utf-8")
    zip_ok = zipfile.is_zipfile(xlsx_path)
    checks.append(Check("workbook_zip", "pass" if zip_ok else "fail", str(xlsx_path)))

    verification = {
        "status": style_status(checks),
        "report_dir": str(report_dir),
        "week_start": str(dates["week_start"]),
        "week_end": str(dates["week_end"]),
        "prior_start": str(dates["prior_start"]),
        "prior_end": str(dates["prior_end"]),
        "outputs": {"xlsx": str(xlsx_path), "html": str(html_path), "verification": str(verification_path)},
        "checks": [asdict(c) for c in checks],
        "metrics": context["metrics"],
    }
    verification_path.write_text(json.dumps(verification, indent=2, default=str), encoding="utf-8")

    print(json.dumps({"status": verification["status"], "xlsx": str(xlsx_path), "html": str(html_path), "verification": str(verification_path)}, indent=2))
    if verification["status"] == "fail":
        return 1
    if args.strict and verification["status"] != "pass":
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
